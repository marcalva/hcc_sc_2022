
import os
import sys
import pandas as pd

from openpyxl import Workbook, utils
from openpyxl.styles import Font, numbers 

from py_excel import *

os.chdir("../../")

fn = "exp/tcga_hcc/ctp.cell_type_main/cell_type_main.marker_genes.txt"
tcga_mrk =  pd.read_csv(fn, sep = '\t', header = 0)

fn = "exp/GSE14520/ctp.cell_type_main/cell_type_main.marker_genes.txt"
lci_mrk = pd.read_csv(fn, sep = '\t', header = 0)


# subset to output data
col_names = ["Main cell-type", "Ensembl ID", "Symbol", "R"]
tcga_mrk.columns = col_names
lci_mrk.columns = col_names

# column name key
key_dict = {"Main cell-type" : "Main cell-type of the marker genes selected by Bisque for decomposition (cell-type proportion estimation)", 
"Ensembl ID" : "Ensembl ID of the marker genes selected by Bisque for decomposition", 
"Symbol" : "Gene symbol of the marker gene used for decomposition", 
"R" : "Pearson correlation coefficient between marker gene expression and estimated cell-type proportion. This gives the contribution and importance of the gene in proportion estimation."
}

key_l = []
for n in col_names:
    key_l.append(key_dict[n])

num_formats = ["@", "@", "@", "0.000"]

# create new workbook
wb = Workbook()

sheet_title = "Supplemental Table 3: Bisque marker genes used for decomposition (for explanations of the columns, please see the table footnote)."

entry_font = Font(size = 12, name = "Calibri")
header_font = Font(size = 12, bold = True, name = "Calibri")

#===============================================================================
# TCGA decomposition markers
#===============================================================================

ws = wb.active 
ws.title = "TCGA"

# header title
title = ws.cell(row=1, column=1)
title.value = sheet_title
title.font = Font(size = 12,
        bold = True,
        name = "Calibri")
title.number_format = "@"

pd2wrksht(tcga_mrk, ws, num_formats, skip = 1, entry_font = entry_font, 
        header_font = header_font)

ws.append([""])
ix = ws.max_row + 1

col1_font = Font(size = 12, bold = True, name = "Calibri")
col2_font = Font(size = 12, bold = False, name = "Calibri")

append_list2wrksht(col_names, ws, row = ix, font = col1_font, col = 1)
append_list2wrksht(key_l, ws, row = ix, font = col2_font, col = 2)

ws.column_dimensions["A"].width = 20

#===============================================================================
# LCI decomposition markers
#===============================================================================

ws = wb.create_sheet("LCI")

# header title
title = ws.cell(row=1, column=1)
title.value = sheet_title
title.font = Font(size = 12,
        bold = True,
        name = "Calibri")
title.number_format = "@"

pd2wrksht(lci_mrk, ws, num_formats, skip = 1, entry_font = entry_font, 
        header_font = header_font)

ws.append([""])
ix = ws.max_row + 1

col1_font = Font(size = 12, bold = True, name = "Calibri")
col2_font = Font(size = 12, bold = False, name = "Calibri")

append_list2wrksht(col_names, ws, row = ix, font = col1_font, col = 1)
append_list2wrksht(key_l, ws, row = ix, font = col2_font, col = 2)

ws.column_dimensions["A"].width = 20

dir_exp = "exp/manuscript/"
os.makedirs(dir_exp, exist_ok = True)

out_fn = dir_exp + "TableS3.bisque_markers.xlsx"
wb.save(out_fn)

