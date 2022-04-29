
import os
import sys
import pandas as pd

from openpyxl import Workbook, utils
from openpyxl.styles import Font, numbers 

from py_excel import *

os.chdir("../../")

# read in mutation results
fn = "exp/tcga_hcc/mol/cell_type_main.SMG_mutsig.2CV.ctp.txt"
mut = pd.read_csv(fn, sep = '\t', header = 0)

# rename columns
col_names = ["W statistic", "W p-value", "T statistic", "Mean mut",
        "Mean WT", "T p-value", "Main cell-type", "Gene", 
        "W adjusted p", "T adjusted p"]
mut.columns = col_names

# Re-order columns
col_k = ["Main cell-type", "Gene", 
        "Mean mut", "Mean WT",
        "W statistic", "W p-value", 
        "T statistic",  "T p-value", 
        "W adjusted p", "T adjusted p"]
mut = mut.loc[:,col_k]

# fonts
hdr_font = Font(size = 12, bold = True, name = "Calibri")
ent_font = Font(size = 12, bold = False, name = "Calibri")

# create new workbook
wb = Workbook()

# worksheet
ws = wb.active 
ws.title = "TCGA"

title = ws.cell(row=1, column=1)
title.value = "Supplemental Table 6: Associations between main cell-type cell-type proportions and somatic mutations in TCGA (for descriptions of the columns, please see the table footnote)."
title.font = hdr_font

# add table
num_formats = ["@", "@", 
"0.00", "0.00", 
"0", "0.00E+00", 
"0.00", "0.00E+00", 
"0.00E+00", "0.00E+00"]

pd2wrksht(mut, ws, num_formats, skip = 1, entry_font = ent_font, 
        header_font = hdr_font, width_mult = 1.5)

ws.append([""])

# key
# ix = ws.max_row + 1
# c = ws.cell(row=ix, column=1, value = "Column")
# c.font = hdr_font
# c = ws.cell(row=ix, column=2, value = "Description")
# c.font = hdr_font
# ix = ix + 1
ix = ws.max_row + 1

append_list2wrksht(mut.columns, ws, row = ix, font = hdr_font, col = 1)

keydict = {
"Main cell-type": "Cell-type tested for proportion differences between patients with mutation (mut) and wildtype (WT)",
"Gene": "The gene for which mutation (mut) and wildtype (WT) are tested for", 
"Mean mut": "Mean of cell-type proportions in individuals who harbor a somatic mutation (mut) \
in the gene", 
"Mean WT": "Mean of cell-type proportions in individuals who have no somatic mutation (WT) \
in the gene", 
"W statistic": "Wilcoxon test statistic for mutation (mut) and wildtype (WT) \
proportion differences", 
"W p-value": "Wilcoxon p-value for for mutation (mut) and wildtype (WT) \
proportion differences", 
"T statistic": "T-test statistic for mutation (mut) and wildtype (WT) \
proportion differences",
"T p-value": "T-test p-value for mutation (mut) and wildtype (WT) \
proportion differences", 
"W adjusted p": "FDR-adjusted Wilcoxon p-value", 
"T adjusted p": "FDR-adjusted T-test p-value"}
l = [keydict[x] for x in col_k]

append_list2wrksht(l, ws, row = ix, font = ent_font, col = 2)

dir_exp = "exp/manuscript/"
os.makedirs(dir_exp, exist_ok = True)

out_fn = dir_exp + "TableS6.mut_ctp.xlsx"
wb.save(out_fn)

