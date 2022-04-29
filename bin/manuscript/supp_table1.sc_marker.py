
import os
import sys
import pandas as pd

from openpyxl import Workbook, utils
from openpyxl.styles import Font, numbers 

from py_excel import *

os.chdir("../../")

fn = "exp/sharma_aiz/markers/markers.cell_type_fine.txt"
fine_mrk = pd.read_csv(fn, sep = '\t', header = 0)

# subset to output data
col_k = ["cluster", "gene", "Name", "p_val", "p_val_adj", 
    "avg_log2FC", "pct.1", "pct.2"]
fine_mrk = fine_mrk.loc[:,col_k]
col_names = ["Subcell-type", "Ensembl ID", "Symbol", "P-value", "Adjusted p-value",
    "Average logFC", "Pct.1", "Pct.2"]
fine_mrk.columns = pd.Index(col_names)

# create new workbook
wb_fine = Workbook()
ws = wb_fine.active 
ws.title = "subtypes"

# header title
title = ws.cell(row=1, column=1)
title.value = "Supplemental Table 1: Cell type marker genes"
title.value = "Supplemental Table 1: Subcell-type marker genes (for explanations of the columns, please see the table footnote)."
title.font = Font(size = 12,
        bold = True,
        name = "Calibri")
title.number_format = "@"

entry_font = Font(size = 12, name = "Calibri")
header_font = Font(size = 12, bold = True, name = "Calibri")
num_formats = ["@", "@", "@", "0.00E+00", "0.00E+00", "0.000", "0.000", "0.000"]

pd2wrksht(fine_mrk, ws, num_formats, skip = 1, entry_font = entry_font, 
        header_font = header_font)

# add column name key
key_dict = {"Subcell-type" : "Subcell-type of marker genes", 
"Ensembl ID" : "Ensembl ID of the marker gene", 
"Symbol" : "Gene symbol of the marker gene", 
"P-value" : "p-value from logistic regression test",
"Adjusted p-value" : "FDR-corrected p-value", 
"Average logFC" : "The log fold change of average expression of droplets in the subcell-type over all others",
"Pct.1" : "Percent of droplets in subcell-type expressing the gene", 
"Pct.2" : "Percent of droplets not in subcell-type expressing the gene"
} 

key_l = []
for n in col_names:
    key_l.append(key_dict[n])

ws.append([""])
ix = ws.max_row + 1

col1_font = Font(size = 12, bold = True, name = "Calibri")
col2_font = Font(size = 12, bold = False, name = "Calibri")

append_list2wrksht(col_names, ws, row = ix, font = col1_font, col = 1)
append_list2wrksht(key_l, ws, row = ix, font = col2_font, col = 2)

ws.column_dimensions["A"].width = 20

dir_exp = "exp/manuscript/"
os.makedirs(dir_exp, exist_ok = True)

out_fn = dir_exp + "TableS1.cell_type_markers.xlsx"
wb_fine.save(out_fn)

