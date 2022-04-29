
import os
import sys
import pandas as pd

from openpyxl import Workbook, utils
from openpyxl.styles import Font, numbers 

from py_excel import *

os.chdir("../../")

# read in enrichment results
fn = "exp/sharma_aiz/path_enr/markers.res.1.gse.Reactome.txt"
enr = pd.read_csv(fn, sep = '\t', header = 0)

fn = "exp/sharma_aiz/clust2ct/res1_to_fine.txt"
ctmap = pd.read_csv(fn, sep = '\t', header = 0, index_col = 0)
ctmap = ctmap['cell_type'].to_dict()

# assign cell types
enr.loc[:,"CellType"] = enr.loc[:,"CellType"].map(ctmap)

# subset columns
col_k = ["CellType", "ID", "Description", "NES", "pvalue", "qvalues", 
        "core_enrichment"]
enr = enr.loc[:,col_k]

# subse rows by q < 0.05
q_k = enr["qvalues"] < 0.05
enr = enr.loc[q_k,:]

# rename columns
col_names = ["Subcell-type", "ID", "Description", "NES", "P-value", "Q-value", 
        "core_enrichment (enrichment genes)"]
enr.columns = col_names

# create new workbook
wb_fine = Workbook()
ws = wb_fine.active 
ws.title = "Reactome"

# header title
title = ws.cell(row=1, column=1)
title.value = "Supplemental Table 2: Reactome gene set enrichment analysis (for explanations of the coulumns, please see the table footnote)."
title.font = Font(size = 12,
        bold = True,
        name = "Calibri")
title.number_format = "@"

entry_font = Font(size = 12, name = "Calibri")
header_font = Font(size = 12, bold = True, name = "Calibri")
num_formats = ["@", "@", "@", "0.000", "0.00E+00", "0.00E+00", "@"]

pd2wrksht(enr, ws, num_formats, skip = 1, entry_font = entry_font, 
        header_font = header_font)

# add column name key
key_dict = {"ID" : "Reactome pathway identifier", 
        "Description" : "Description of the pathway", 
        "NES" : "Normalized enrichment score from GSEA", 
        "P-value" : "P value from GSEA", 
        "Q-value" : "FDR-adjusted p-value from GSEA", 
        "core_enrichment (enrichment genes)" : "Genes contributing to enrichment", 
        "Subcell-type" : "Subcell-type of the ranked gene list used for GSEA"}

key_l = []
for n in col_names:
    key_l.append(key_dict[n])

ws.append([""])
ix = ws.max_row + 1

col1_font = Font(size = 12, bold = True, name = "Calibri")
col2_font = Font(size = 12, bold = False, name = "Calibri")

append_list2wrksht(col_names, ws, row = ix, font = col1_font, col = 1)
append_list2wrksht(key_l, ws, row = ix, font = col2_font, col = 2)

ws.column_dimensions["A"].width = 20

dir_exp = "exp/manuscript/"
os.makedirs(dir_exp, exist_ok = True)

out_fn = dir_exp + "TableS2.Reactome.xlsx"
wb_fine.save(out_fn)

