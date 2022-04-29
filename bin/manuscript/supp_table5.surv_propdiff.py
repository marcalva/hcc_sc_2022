
import os
import sys
import pandas as pd

from openpyxl import Workbook, utils
from openpyxl.styles import Font, numbers 

from py_excel import *

os.chdir("../../")

# read in diff results
fn = "exp/tcga_hcc/survival/cell_type_main.cox_ph.formatted.txt"
surv = pd.read_csv(fn, sep = '\t', header = 0)

# rename columns
col_names = ["Event", "Model", "Main cell-type", "N", "HR", 
"95% CI", "P-value", "Adjusted p-value"]
surv.columns = col_names

# fonts
hdr_font = Font(size = 12, bold = True, name = "Calibri")
ent_font = Font(size = 12, bold = False, name = "Calibri")

# create new workbook
wb = Workbook()

# key sheet
ws = wb.active 
ws.title = "TCGA"

title = ws.cell(row=1, column=1)
title.value = "Supplemental Table 5: Association of main cell-type proportions with survival outcomes in TCGA"
title.value = "Supplemental Table 5: Associations of main cell-type proportions with survival outcomes in TCGA (for descriptions of the columns, please see the table footnote)."
title.font = hdr_font

# add table
num_formats = ["@", "@", "@", "0", "0.00", "@", "@", "@"]

pd2wrksht(surv, ws, num_formats, skip = 1, entry_font = ent_font, 
        header_font = hdr_font, width_mult = 1.5)

# add key
ws.append([""])

ix = ws.max_row + 1

append_list2wrksht(col_names, ws, row = ix, font = hdr_font, col = 1)

keydict = {
"Event": "Survival outcome", 
"Model": "Cox proportional hazards regression model", 
"Main cell-type": "Main cell-type for which proportions are tested for", 
"N": "Total number of individuals used in the survival outcome analysis", 
"HR": "Hazard ratio from Cox proportional hazards regression model", 
"95% CI": "95% confidence interval from Cox proportional hazards regression model", 
"P-value": "P-value from Cox proportional hazards regression model", 
"Adjusted p-value": "FDR-adjusted p-value from Cox proportional hazards regression model"
}
l = [keydict[x] for x in col_names]

append_list2wrksht(l, ws, row = ix, font = ent_font, col = 2)

ix = ws.max_row + 1
append_list2wrksht(["NS"], ws, row = ix, font = hdr_font, col = 1)
append_list2wrksht(["P-value or adjusted p-value is non-significant (P=>0.05)"], 
    ws, row = ix, font = ent_font, col = 2)

ws.column_dimensions["A"].width = 20

dir_exp = "exp/manuscript/"
os.makedirs(dir_exp, exist_ok = True)

out_fn = dir_exp + "TableS5.surv.xlsx"
wb.save(out_fn)

