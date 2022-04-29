
def fmt_width_exc(s):
    import re
    if re.search("^0\.0+$", s):
        return True
    if re.search("^0\.0+E\+0+$", s):
        return True
    return False

def pd2wrksht(datf, wrksht, formats, skip=0, entry_font = -1, header=True,
        header_font = -1, rowname = -1, rowname_font = -1, width_mult = 1.2):
    from openpyxl import Workbook, utils
    from openpyxl.styles import Font, numbers
    import re
    if entry_font == -1:
        entry_font = Font()
    if header:
        col_names = datf.columns.to_list()
        if header_font == -1:
            header_font = Font()
        for n in range(len(col_names)):
            c = wrksht.cell(row = skip + 1, column = n + 1, value = col_names[n])
            c.font = header_font
            c.number_format = "@"
        skip = skip + 1
    for i in range(datf.shape[0]):
        for j in range(datf.shape[1]):
            val = datf.iloc[i,j]
            c = wrksht.cell(row = i + 1 + skip, column = j + 1, value = val)
            c.font = entry_font
            c.number_format = formats[j]
    if rowname > -1:
        if rowname_font == -1:
            rowname_font = Font()
        for i in range(datf.shape[0]):
            c = wrksht.cell(row = i + 1 + skip, column = rowname + 1, value = val)
    # set column widths
    for i, col_cells in enumerate(wrksht.columns):
        if fmt_width_exc(formats[i]): # length of format ID (eg 0.00E+00)
            ss = [formats[i]]
            if header:
                ss.append(col_cells[skip-1].value)
            new_len = max(len(str(s)) for s in ss)
        else:
            ss = col_cells[(skip-1):]
            new_len = max(len(str(s.value)) for s in ss)
        col_let = utils.get_column_letter(col_cells[0].column)
        wrksht.column_dimensions[col_let].width = new_len * width_mult
    return None

def append_list2wrksht(l, wrksht, row = -1, font = -1, num_format = "General", col = 1):
    from openpyxl import Workbook, utils
    from openpyxl.styles import Font, numbers
    if font == -1:
        font = Font()
    if row < 0:
        row = wrksht.max_row + 1
    for i in range(len(l)):
        c = wrksht.cell(row = row + i, column = col, value = l[i])
        c.font = font
        c.number_format = num_format

