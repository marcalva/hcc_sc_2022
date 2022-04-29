
import os
import sys
import pandas as pd

from openpyxl import Workbook, utils
from openpyxl.styles import Font, numbers 

from py_excel import *

os.chdir("../../")

# read in diff results
fn = "exp/tcga_hcc/ctp.cell_type_main/tcga.cell_type_main.tum_nontum.txt"
tcga_tumdif = pd.read_csv(fn, sep = '\t', header = 0, index_col = 0)
tcga_tumdif.loc[:,"Cohort"] = "TCGA"

fn = "exp/GSE14520/ctp.cell_type_main/lci.cell_type_main.tum_nontum.txt"
lci_tumdif = pd.read_csv(fn, sep = '\t', header = 0, index_col = 0)
lci_tumdif.loc[:,"Cohort"] = "LCI"

# rename columns
col_names = ["W statistic", "W p-value", "T statistic", "T estimate", 
"T p-value", "W adjusted p", "T adjusted p", "Main cell-type", "Cohort"]
tcga_tumdif.columns = col_names
lci_tumdif.columns = col_names

k = lci_tumdif["Main cell-type"] != "B"
lci_tumdif = lci_tumdif[k]

tumdif = pd.concat([tcga_tumdif, lci_tumdif], axis = 0)

# Re-order columns
col_k = ["Cohort", "Main cell-type", "W statistic", "W p-value", 
        "T statistic", "T estimate", "T p-value", 
        "W adjusted p", "T adjusted p"]
tumdif = tumdif[col_k]

# fonts
hdr_font = Font(size = 12, bold = True, name = "Calibri")
ent_font = Font(size = 12, bold = False, name = "Calibri")

# create new workbook
wb = Workbook()

# key sheet
ws = wb.active 
ws.title = "Tumor proportions"

title = ws.cell(row=1, column=1)
title.value = "Supplemental Table 4. Proportions of the Prol cell-type differ most between the HCC tumor and adjacent non-tumor tissue in the TCGA and LCI bulk tissue cohorts (for descriptions of the columns, please see the table footnote)."
title.font = hdr_font

# column formats
num_formats = ["@", "@", "0.00", "0.00E+00", "0.00", "0.00", "0.00E+00", 
        "0.00E+00", "0.00E+00"]

pd2wrksht(tumdif, ws, num_formats, entry_font = ent_font, 
        header_font = hdr_font, skip = 1)

ws.append([""])

ix = ws.max_row + 1

append_list2wrksht(col_k, ws, row = ix, font = hdr_font, col = 1)

keydict = {
"Cohort": "TCGA (n=49) or LCI (n=209) cohort", 
"Main cell-type": "Main cell-type classification", 
"W statistic": "Paired Wilcoxon test statistic between tumor and \
adjacent non-tumor proportion estimates", 
"W p-value": "Paired Wilcoxon test p-value between tumor and \
adjacent non-tumor proportion estimates", 
"T statistic": "Paired T-test statistic between tumor and \
adjacent non-tumor proportion estimates", 
"T estimate": "Paired T-test estimate between tumor and adjacent \
non-tumor proportion estimates",
"T p-value": "Paired T-test p-value between tumor and adjacent non-tumor \
proportion estimates", 
"W adjusted p": "FDR-adjusted p-value for paired Wilcoxon test", 
"T adjusted p": "FDR-adjusted p-value for paired T-test"
        }
l = [keydict[x] for x in col_k]

append_list2wrksht(l, ws, row = ix, font = ent_font, col = 2)

# set widths
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 20

dir_exp = "exp/manuscript/"
os.makedirs(dir_exp, exist_ok = True)

out_fn = dir_exp + "TableS4.tum_ctpdiff.xlsx"
wb.save(out_fn)

